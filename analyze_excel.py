import openpyxl
import sys

# Ouvrir le fichier Excel
filename = sys.argv[1] if len(sys.argv) > 1 else "donnees1.xlsx"
wb = openpyxl.load_workbook(filename)

print(f"\n📊 Analyse de {filename}")
print(f"=" * 60)

# Lister toutes les feuilles
print(f"\n📋 Feuilles disponibles: {wb.sheetnames}")

# Analyser chaque feuille
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    print(f"\n📄 Feuille: {sheet_name}")
    print(f"-" * 60)

    # Compter le nombre de lignes avec données
    data_rows = 0
    for row in ws.iter_rows(min_row=2):  # Skip header
        if row[0].value is not None:
            data_rows += 1

    print(f"Nombre de points de données: {data_rows}")

    if data_rows > 0:
        # Afficher les 10 premières lignes
        print(f"\n🔝 Premières valeurs:")
        print(f"{'Episode':<12} {'Score':<12}")
        print("-" * 24)
        count = 0
        for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
            if row[0] is not None:
                print(f"{row[0]:<12} {row[1]:<12.4f}")
                count += 1

        # Afficher les 10 dernières lignes
        if data_rows > 10:
            print(f"\n🔚 Dernières valeurs:")
            print(f"{'Episode':<12} {'Score':<12}")
            print("-" * 24)
            all_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    all_data.append(row)

            for row in all_data[-10:]:
                print(f"{row[0]:<12} {row[1]:<12.4f}")

        # Calculer statistiques
        scores = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                scores.append(row[1])

        if scores:
            print(f"\n📈 Statistiques:")
            print(f"Score minimum:  {min(scores):.4f}")
            print(f"Score maximum:  {max(scores):.4f}")
            print(f"Score moyen:    {sum(scores)/len(scores):.4f}")

            # Tendance: comparer première moitié vs deuxième moitié
            mid = len(scores) // 2
            if mid > 0:
                first_half_avg = sum(scores[:mid]) / mid
                second_half_avg = sum(scores[mid:]) / (len(scores) - mid)
                improvement = second_half_avg - first_half_avg

                print(f"\n📊 Progression:")
                print(f"Moyenne 1ère moitié:  {first_half_avg:.4f}")
                print(f"Moyenne 2ème moitié:  {second_half_avg:.4f}")
                print(f"Amélioration:         {improvement:+.4f}")

                if improvement > 0.5:
                    print("✅ L'IA APPREND! Progression positive claire!")
                elif improvement > 0.1:
                    print("🟡 Légère amélioration, mais lente")
                elif improvement > -0.1:
                    print("⚠️  Stagnation, pas de progression claire")
                else:
                    print("❌ Régression, l'IA n'apprend pas")

print("\n" + "=" * 60)
