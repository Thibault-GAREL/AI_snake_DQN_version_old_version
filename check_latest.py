import openpyxl
import sys

# Ouvrir le fichier Excel
filename = sys.argv[1] if len(sys.argv) > 1 else "donnees2.xlsx"
wb = openpyxl.load_workbook(filename)

print(f"\n📊 Feuilles dans {filename}:")
print("=" * 60)

sheets_info = []

for sheet_name in wb.sheetnames:
    if sheet_name == "Sheet":
        continue

    ws = wb[sheet_name]

    # Compter le nombre de lignes avec données
    data_rows = 0
    last_episode = None
    last_score = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data_rows += 1
            last_episode = row[0]
            last_score = row[1]

    if data_rows > 0:
        sheets_info.append({
            'name': sheet_name,
            'rows': data_rows,
            'last_episode': last_episode,
            'last_score': last_score
        })

# Afficher triées par numéro d'entraînement
sheets_info.sort(key=lambda x: x['name'])

for info in sheets_info:
    print(f"\n📄 {info['name']}")
    print(f"   Points de données: {info['rows']}")
    print(f"   Dernier épisode: {info['last_episode']}")
    print(f"   Score final: {info['last_score']:.4f}")

print("\n" + "=" * 60)
print(f"\n🎯 DERNIÈRE FEUILLE: {sheets_info[-1]['name']}")
print(f"   Score: {sheets_info[-1]['last_score']:.4f} après {sheets_info[-1]['last_episode']} épisodes")
print("=" * 60)
